import threading
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox
import openpyxl
import requests
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from io import BytesIO


class ImageColumnApp:

    def __init__(self, root):
        self.skip = 0

        self.root = root
        self.root.title("REPLENISH-IMAGE")

        root.geometry(f"{200}x{250}")

        self.file_path = tk.StringVar()
        self.column_name = tk.StringVar()

        self.select_xlsx_file_label = tk.Label(root, text="Select xlsx file:")
        self.select_xlsx_file_label.pack()

        self.xlsx_file_path = tk.Entry(root, textvariable=self.file_path)
        self.xlsx_file_path.pack()

        self.browse_button = tk.Button(root, text="Browse", command=self.browse_file)
        self.browse_button.pack()

        self.enter_column_name_label = tk.Label(root, text="Enter column name:")
        self.enter_column_name_label.pack()

        self.column_name_input = tk.Entry(root, textvariable=self.column_name)
        self.column_name_input.pack()

        self.process_button = tk.Button(root, text="Process", command=start_process_thread)
        self.process_button.pack()

        self.process_label = tk.StringVar()
        self.process_label.set("Ready")

        self.label3 = tk.Label(root, textvariable=self.process_label)
        self.label3.pack()

        self.skip_label = tk.StringVar()
        self.skip_label.set("")

        self.label4 = tk.Label(root, textvariable=self.skip_label)
        self.label4.pack()

    def browse_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.file_path.set(file_path)

    def download_image(self, url):
        try:
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                new_image = PILImage.open(BytesIO(response.content))
                if 'WEBP' in new_image.format:
                    new_image = new_image.convert("RGB")
                    img_data = BytesIO()
                    new_image.save(img_data, format="JPEG")
                    return Image(img_data)
                elif 'MPO' in new_image.format:
                    return None
                return Image(new_image)
        except Exception as e:
            print(str(e))
            self.skip = self.skip + 1
            pass
        return None

    def get_column_index(self):
        file_path = self.file_path.get()
        column_name = self.column_name.get()

        if file_path and column_name:
            try:
                workbook = load_workbook(file_path)
                sheet = workbook.active

                max_column = sheet.max_column
                for column in range(1, max_column + 1):
                    if column_name.strip() == sheet.cell(row=1, column=column).value:
                        return column

            except Exception as e:
                messagebox.showerror("Error", str(e))

    def process_excel(self):
        self.skip = 0
        file_path = self.file_path.get()
        column = self.get_column_index()
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            max_row = sheet.max_row

            image_column = sheet.max_column + 1

            # Inside the loop where you add images:
            for row in range(2, max_row + 1):
                self.process_label.set("Processing, row number: " + str(row))
                self.skip_label.set("Skip row number: " + str(self.skip))
                try:
                    url_cell = sheet.cell(row=row, column=column)
                    image_url = url_cell.value
                    if image_url and image_url.startswith("http"):
                        img = self.download_image(image_url)
                        if img:
                            img_cell = sheet.cell(row=row, column=image_column)
                            sheet.row_dimensions[row].height = 120
                            sheet.column_dimensions[get_column_letter(img_cell.column)].width = 20
                            img.width = 150
                            img.height = 150
                            sheet.add_image(img, img_cell.coordinate)
                        else:
                            for cell in sheet[row]:
                                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

                except Exception as e:
                    print("Error:", str(e))  # Print error to console
                    pass  # Continue with the loop

            current_time = datetime.now().strftime("%Y%m%d%H%M%S")
            new_file_path = file_path.replace(".xlsx", f"_{current_time}_with_images.xlsx")

            workbook.save(new_file_path)
            workbook.close()

            messagebox.showinfo("Success", "Images added to the new column!")
            self.process_label.set("Success!")
        except Exception as e:
            print("Error:", str(e))
            messagebox.showerror("Error", str(e))
            self.process_label.set("Fail!")


def start_process_thread():
    def process_thread():
        try:
            app.process_excel()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    thread = threading.Thread(target=process_thread)
    thread.start()


if __name__ == '__main__':
    root = tk.Tk()
    app = ImageColumnApp(root)
    root.mainloop()
